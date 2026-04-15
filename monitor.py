#!/usr/bin/env python3
"""
arxiv LLM 量化论文自动监控主脚本
每天定时执行：搜索 -> 查重 -> 下载 PDF -> 输出结构化 JSON
中文总结和作者单位提取由 hermes cronjob agent 调用 LLM 完成
"""

import os
import sys
import json
import time
import xml.etree.ElementTree as ET
from datetime import datetime, date
from pathlib import Path

import requests
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# ==================== 配置区 ====================
BASE_DIR = Path(__file__).resolve().parent
PAPERS_DIR = BASE_DIR / "papers"
EXCEL_FILE = BASE_DIR / "papers_record.xlsx"
VIEWER_JSON = BASE_DIR / "viewer" / "papers_data.json"
CRAWLED_IDS_FILE = BASE_DIR / "crawled_ids.txt"
PENDING_LLM_IDS_FILE = BASE_DIR / "pending_llm_ids.txt"
KEYWORDS_FILE = BASE_DIR / "search_keywords.txt"
OUTPUT_JSON = BASE_DIR / "new_papers.json"   # 输出给 hermes agent 的中间文件

# arxiv API 配置
ARXIV_API = "https://export.arxiv.org/api/query"
MAX_RESULTS = 50
REQUEST_INTERVAL = 3  # 秒

# ==================== 工具函数 ====================

def load_crawled_ids() -> set:
    if not CRAWLED_IDS_FILE.exists():
        return set()
    with open(CRAWLED_IDS_FILE, "r", encoding="utf-8") as f:
        return set(line.strip() for line in f if line.strip())


def load_excel_ids() -> set:
    """从 Excel 的 Papers 表读取已有 arxiv_id。"""
    if not EXCEL_FILE.exists():
        return set()
    try:
        wb = openpyxl.load_workbook(EXCEL_FILE, read_only=True)
        if "Papers" not in wb.sheetnames:
            return set()
        ws = wb["Papers"]
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not header_row:
            return set()
        headers = [str(h) if h is not None else "" for h in header_row]
        if "arxiv_id" not in headers:
            return set()
        arxiv_col = headers.index("arxiv_id")
        ids = set()
        for row in ws.iter_rows(min_row=2, values_only=True):
            val = row[arxiv_col] if arxiv_col < len(row) else None
            if val:
                ids.add(str(val).strip())
        return ids
    except Exception as e:
        print(f"[WARN] Failed to load Excel IDs: {e}")
        return set()


def load_pending_llm_ids() -> set:
    if not PENDING_LLM_IDS_FILE.exists():
        return set()
    with open(PENDING_LLM_IDS_FILE, "r", encoding="utf-8") as f:
        return set(line.strip() for line in f if line.strip())


def save_pending_llm_ids(ids: set[str] | list[str]):
    cleaned = sorted({str(x).strip() for x in ids if str(x).strip()})
    if not cleaned:
        if PENDING_LLM_IDS_FILE.exists():
            PENDING_LLM_IDS_FILE.unlink()
        return
    with open(PENDING_LLM_IDS_FILE, "w", encoding="utf-8") as f:
        for arxiv_id in cleaned:
            f.write(arxiv_id + "\n")


def save_crawled_ids_batch(new_ids: list[str]):
    """批量追加新 ID"""
    with open(CRAWLED_IDS_FILE, "a", encoding="utf-8") as f:
        for arxiv_id in new_ids:
            f.write(arxiv_id + "\n")


def load_search_keywords() -> str:
    default_keywords = "all:quantization+AND+all:large+AND+all:language+AND+all:model"
    if KEYWORDS_FILE.exists():
        with open(KEYWORDS_FILE, "r", encoding="utf-8") as f:
            kw = f.read().strip()
            return kw if kw else default_keywords
    return default_keywords


def search_arxiv_papers(keywords: str, max_results: int = MAX_RESULTS) -> list[dict]:
    url = (
        f"{ARXIV_API}?search_query={keywords}"
        f"&max_results={max_results}"
        f"&sortBy=submittedDate"
        f"&sortOrder=descending"
    )

    print(f"[INFO] Searching arxiv: {keywords}")
    response = requests.get(url, timeout=30)
    response.raise_for_status()

    ns = {"a": "http://www.w3.org/2005/Atom"}
    root = ET.fromstring(response.content)

    papers = []
    for entry in root.findall("a:entry", ns):
        try:
            arxiv_id_full = entry.find("a:id", ns).text.strip().split("/abs/")[-1]
            arxiv_id = arxiv_id_full.split("v")[0]   # 剥离版本号

            title = entry.find("a:title", ns).text.strip().replace("\n", " ")
            authors = ", ".join(
                a.find("a:name", ns).text for a in entry.findall("a:author", ns)
            )
            summary = entry.find("a:summary", ns).text.strip().replace("\n", " ")
            published = entry.find("a:published", ns).text[:10]
            cats = ", ".join(c.get("term") for c in entry.findall("a:category", ns))
            pdf_url = f"https://arxiv.org/pdf/{arxiv_id_full}"

            papers.append({
                "arxiv_id": arxiv_id,
                "title": title,
                "authors": authors,
                "summary": summary,
                "published_date": published,
                "categories": cats,
                "pdf_url": pdf_url,
                "pdf_filename": f"{arxiv_id}.pdf",
                "pdf_local_path": str(PAPERS_DIR / f"{arxiv_id}.pdf"),
                "affiliations": "",   # ← 由 hermes LLM 从 PDF 提取
                "summary_cn": "",      # ← 由 hermes LLM 生成中文总结
            })
        except Exception as e:
            print(f"[WARN] Parse error: {e}")
            continue

    return papers


def download_pdf(paper: dict) -> bool:
    pdf_path = PAPERS_DIR / paper["pdf_filename"]
    if pdf_path.exists():
        print(f"[INFO] PDF exists: {paper['pdf_filename']}")
        return True
    try:
        response = requests.get(paper["pdf_url"], timeout=60, stream=True)
        response.raise_for_status()
        with open(pdf_path, "wb") as f:
            for chunk in response.iter_content(chunk_size=8192):
                f.write(chunk)
        print(f"[INFO] Downloaded: {paper['pdf_filename']}")
        return True
    except Exception as e:
        print(f"[ERROR] Download failed {paper['pdf_url']}: {e}")
        return False


def load_or_create_excel() -> openpyxl.Workbook:
    if EXCEL_FILE.exists():
        wb = openpyxl.load_workbook(EXCEL_FILE)
        if "Papers" not in wb.sheetnames:
            wb.create_sheet("Papers")
        return wb

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Papers"
    headers = [
        "arxiv_id", "title", "authors", "affiliations",
        "published_date", "categories", "abstract", "summary_cn",
        "pdf_filename", "crawled_date", "notes"
    ]
    ws.append(headers)

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for col, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    col_widths = [15, 45, 28, 35, 12, 18, 70, 60, 20, 12, 25]
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

    return wb


def append_to_excel(wb: openpyxl.Workbook, paper: dict):
    ws = wb["Papers"]
    today = date.today().isoformat()
    row = [
        paper["arxiv_id"],
        paper["title"],
        paper["authors"],
        paper.get("affiliations", ""),
        paper["published_date"],
        paper["categories"],
        paper["summary"],
        paper.get("summary_cn", ""),
        paper["pdf_filename"],
        today,
        "",  # notes
    ]
    ws.append(row)
    last_row = ws.max_row
    for col in range(1, len(row) + 1):
        ws.cell(row=last_row, column=col).alignment = Alignment(wrap_text=True, vertical="top")
    print(f"[INFO] Appended: {paper['arxiv_id']} - {paper['title'][:40]}...")


def build_excel_row_index(ws: openpyxl.worksheet.worksheet.Worksheet) -> tuple[dict[str, int], dict[str, int]]:
    """返回 (header_index, arxiv_id -> row_number)。"""
    header_row = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    header_index = {str(v): i + 1 for i, v in enumerate(header_row) if v is not None}

    row_index: dict[str, int] = {}
    arxiv_col = header_index.get("arxiv_id")
    if not arxiv_col:
        return header_index, row_index

    for r in range(2, ws.max_row + 1):
        val = ws.cell(row=r, column=arxiv_col).value
        if val is None:
            continue
        key = str(val).strip()
        if key and key not in row_index:
            row_index[key] = r
    return header_index, row_index


def upsert_to_excel(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    header_index: dict[str, int],
    row_index: dict[str, int],
    paper: dict,
):
    """按 arxiv_id 更新或插入，避免重复行。"""
    arxiv_id = paper["arxiv_id"]
    today = date.today().isoformat()

    if arxiv_id in row_index:
        target_row = row_index[arxiv_id]
        updates = {
            "title": paper["title"],
            "authors": paper["authors"],
            "published_date": paper["published_date"],
            "categories": paper["categories"],
            "abstract": paper["summary"],
            "pdf_filename": paper["pdf_filename"],
            "crawled_date": today,
        }
        # 仅在新值非空时覆盖，避免把已有结果清空。
        if paper.get("affiliations"):
            updates["affiliations"] = paper.get("affiliations", "")
        if paper.get("summary_cn"):
            updates["summary_cn"] = paper.get("summary_cn", "")

        for key, value in updates.items():
            col = header_index.get(key)
            if col:
                ws.cell(row=target_row, column=col, value=value)

        for col in range(1, ws.max_column + 1):
            ws.cell(row=target_row, column=col).alignment = Alignment(wrap_text=True, vertical="top")
        print(f"[INFO] Updated: {arxiv_id} - {paper['title'][:40]}...")
        return

    # 不存在则新增
    append_to_excel(wb=ws.parent, paper=paper)
    row_index[arxiv_id] = ws.max_row


def save_excel(wb: openpyxl.Workbook):
    EXCEL_FILE.parent.mkdir(parents=True, exist_ok=True)
    wb.save(EXCEL_FILE)
    print(f"[INFO] Excel saved: {EXCEL_FILE}")


def export_viewer_json_from_excel():
    """从 papers_record.xlsx 导出 viewer 使用的 papers_data.json。"""
    if not EXCEL_FILE.exists():
        print(f"[WARN] Excel not found, skip viewer export: {EXCEL_FILE}")
        return

    wb = openpyxl.load_workbook(EXCEL_FILE, read_only=True)
    if "Papers" not in wb.sheetnames:
        print("[WARN] Sheet 'Papers' not found, skip viewer export")
        return

    ws = wb["Papers"]
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not header_row:
        print("[WARN] Excel header missing, skip viewer export")
        return

    headers = [str(h) if h is not None else "" for h in header_row]
    index = {name: i for i, name in enumerate(headers)}
    required = [
        "arxiv_id",
        "title",
        "authors",
        "affiliations",
        "published_date",
        "categories",
        "abstract",
        "summary_cn",
        "pdf_filename",
        "crawled_date",
        "notes",
    ]
    missing = [c for c in required if c not in index]
    if missing:
        print(f"[WARN] Missing columns in Excel, skip viewer export: {missing}")
        return

    def norm(v: object) -> str:
        if v is None:
            return ""
        return str(v).replace("\n", " ").strip()

    def quality_key(p: dict) -> tuple:
        # 优先保留信息更完整的一行：有中文总结、有单位、文本更长、日期更新。
        return (
            1 if p.get("summary_cn") else 0,
            1 if p.get("affiliations") else 0,
            len(p.get("summary_cn", "")),
            len(p.get("affiliations", "")),
            len(p.get("abstract", "")),
            p.get("crawled_date", ""),
            p.get("published_date", ""),
        )

    papers_by_id: dict[str, dict] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        paper = {col: norm(row[index[col]]) for col in required}
        if not paper["arxiv_id"]:
            continue
        paper["pdf_url"] = f"https://arxiv.org/pdf/{paper['arxiv_id']}"
        arxiv_id = paper["arxiv_id"]
        old = papers_by_id.get(arxiv_id)
        if old is None or quality_key(paper) > quality_key(old):
            papers_by_id[arxiv_id] = paper

    papers = list(papers_by_id.values())

    papers.sort(
        key=lambda x: (x["crawled_date"], x["published_date"], x["arxiv_id"]),
        reverse=True,
    )

    crawled_dates = sorted({p["crawled_date"] for p in papers if p["crawled_date"]})
    published_dates = sorted({p["published_date"] for p in papers if p["published_date"]})

    payload = {
        "count": len(papers),
        "crawled_date_min": crawled_dates[0] if crawled_dates else "",
        "crawled_date_max": crawled_dates[-1] if crawled_dates else "",
        "published_date_min": published_dates[0] if published_dates else "",
        "published_date_max": published_dates[-1] if published_dates else "",
        "papers": papers,
    }

    VIEWER_JSON.parent.mkdir(parents=True, exist_ok=True)
    with open(VIEWER_JSON, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)
    print(f"[INFO] Viewer JSON updated: {VIEWER_JSON} (count={len(papers)})")


def load_incomplete_papers_from_excel() -> dict[str, dict]:
    """读取 Excel 中尚未完成 LLM 补全的论文。"""
    if not EXCEL_FILE.exists():
        return {}

    wb = openpyxl.load_workbook(EXCEL_FILE, read_only=True)
    if "Papers" not in wb.sheetnames:
        return {}

    ws = wb["Papers"]
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not header_row:
        return {}

    headers = [str(h) if h is not None else "" for h in header_row]
    index = {name: i for i, name in enumerate(headers)}
    required = [
        "arxiv_id",
        "title",
        "authors",
        "affiliations",
        "published_date",
        "categories",
        "abstract",
        "summary_cn",
        "pdf_filename",
        "crawled_date",
        "notes",
    ]
    missing = [c for c in required if c not in index]
    if missing:
        print(f"[WARN] Missing columns in Excel, skip pending check: {missing}")
        return {}

    def norm(v: object) -> str:
        if v is None:
            return ""
        return str(v).replace("\n", " ").strip()

    pending: dict[str, dict] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        paper = {col: norm(row[index[col]]) for col in required}
        arxiv_id = paper["arxiv_id"]
        if not arxiv_id:
            continue
        if paper["affiliations"] and paper["summary_cn"]:
            continue
        paper["summary"] = paper["abstract"]
        paper["pdf_url"] = f"https://arxiv.org/pdf/{arxiv_id}"
        paper["pdf_local_path"] = str(PAPERS_DIR / paper["pdf_filename"]) if paper["pdf_filename"] else ""
        pending[arxiv_id] = paper
    return pending


def write_llm_output_json(
    papers_to_process: list[dict],
    fresh_downloaded_count: int = 0,
    feishu_msg: str = "",
):
    """输出当前待处理状态，供 Hermes agent 继续执行或安全重试。"""
    output = {
        "date": date.today().isoformat(),
        "new_count": fresh_downloaded_count,
        "pending_count": len(papers_to_process),
        "excel_file": str(EXCEL_FILE),
        "papers_dir": str(PAPERS_DIR),
        "new_papers": papers_to_process,
        "papers_to_process": papers_to_process,
        "feishu_msg": feishu_msg,
    }
    with open(OUTPUT_JSON, "w", encoding="utf-8") as f:
        json.dump(output, f, ensure_ascii=False, indent=2)


def sync_pending_state_from_excel(refresh_output_json: bool = True) -> list[dict]:
    """
    根据 Excel 当前状态重建 pending_llm_ids.txt。
    可选同时刷新 new_papers.json，避免 agent 重试时继续读取旧待处理列表。
    """
    incomplete_excel_papers = load_incomplete_papers_from_excel()
    papers_to_process = [
        incomplete_excel_papers[arxiv_id]
        for arxiv_id in sorted(incomplete_excel_papers)
    ]
    save_pending_llm_ids({p["arxiv_id"] for p in papers_to_process})
    if refresh_output_json:
        write_llm_output_json(papers_to_process=papers_to_process)
    return papers_to_process


# ==================== 主流程 ====================

def main():
    if len(sys.argv) > 1 and sys.argv[1] == "--sync-pending-state":
        papers_to_process = sync_pending_state_from_excel(refresh_output_json=True)
        print(
            f"[INFO] Pending LLM state synced from Excel | "
            f"remaining={len(papers_to_process)} output_json={OUTPUT_JSON}"
        )
        return

    print("=" * 60)
    print(f"[START] arxiv Monitor | {datetime.now().isoformat()}")
    print("=" * 60)

    PAPERS_DIR.mkdir(parents=True, exist_ok=True)

    # 加载已爬取 ID（查重）：Excel 为主，txt 为兜底缓存。
    crawled_ids_txt = load_crawled_ids()
    crawled_ids_excel = load_excel_ids()
    crawled_ids = crawled_ids_txt | crawled_ids_excel
    pending_ids_file = load_pending_llm_ids()
    incomplete_excel_papers = load_incomplete_papers_from_excel()
    pending_ids = pending_ids_file | set(incomplete_excel_papers.keys())
    save_pending_llm_ids(pending_ids)
    print(
        f"[INFO] crawled IDs loaded | txt={len(crawled_ids_txt)} "
        f"excel={len(crawled_ids_excel)} merged={len(crawled_ids)}"
    )
    print(
        f"[INFO] pending LLM IDs loaded | file={len(pending_ids_file)} "
        f"excel_incomplete={len(incomplete_excel_papers)} merged={len(pending_ids)}"
    )

    # 搜索
    keywords = load_search_keywords()
    all_papers = search_arxiv_papers(keywords)
    print(f"[INFO] Retrieved {len(all_papers)} papers from arxiv")

    # 查重
    new_papers = [p for p in all_papers if p["arxiv_id"] not in crawled_ids]
    print(f"[INFO] {len(new_papers)} NEW papers")

    # 下载 PDF + 更新 ID
    downloaded = []
    for paper in new_papers:
        ok = download_pdf(paper)
        downloaded.append({**paper, "pdf_downloaded": ok})
        time.sleep(REQUEST_INTERVAL)

    if downloaded:
        # 保存 Excel（summary_cn 和 affiliations 暂留空，等 LLM 填入）
        wb = load_or_create_excel()
        ws = wb["Papers"]
        header_index, row_index = build_excel_row_index(ws)
        for paper in downloaded:
            upsert_to_excel(ws, header_index, row_index, paper)
        save_excel(wb)
        export_viewer_json_from_excel()

        # 批量写入 crawled_ids
        save_crawled_ids_batch([p["arxiv_id"] for p in downloaded])
        pending_ids |= {p["arxiv_id"] for p in downloaded}

    incomplete_excel_papers = load_incomplete_papers_from_excel()
    papers_to_process = [incomplete_excel_papers[arxiv_id] for arxiv_id in sorted(pending_ids) if arxiv_id in incomplete_excel_papers]
    unresolved_ids = {p["arxiv_id"] for p in papers_to_process}
    save_pending_llm_ids(unresolved_ids)

    if not papers_to_process:
        # 无新论文，且无待补全论文
        output = {
            "date": date.today().isoformat(),
            "new_count": 0,
            "pending_count": 0,
            "new_papers": [],
            "papers_to_process": [],
            "feishu_msg": f"✅ 今日（{date.today().isoformat()}）未发现新的 LLM 量化论文。",
        }
        with open(OUTPUT_JSON, "w", encoding="utf-8") as f:
            json.dump(output, f, ensure_ascii=False, indent=2)
        export_viewer_json_from_excel()
        print("[INFO] No new papers and no pending LLM tasks. Output JSON written.")
        return

    # 输出 JSON（供 hermes agent 读取并做 LLM summarization）
    write_llm_output_json(
        papers_to_process=papers_to_process,
        fresh_downloaded_count=len(downloaded),
    )

    print(f"[INFO] Output JSON: {OUTPUT_JSON}")
    print(
        f"[INFO] fresh_downloaded={len(downloaded)} "
        f"pending_llm={len(papers_to_process)}. Awaiting LLM summarization..."
    )

    print("\n" + "=" * 60)
    print("[LLM_SUMMARIZATION_REQUIRED]")
    print("=" * 60)
    print(f"JSON file: {OUTPUT_JSON}")
    print(f"Fresh downloads: {len(downloaded)}")
    print(f"Papers awaiting LLM completion: {len(papers_to_process)}")
    for p in papers_to_process:
        print(f"  - [{p['arxiv_id']}] {p['title'][:50]}... | PDF: {p['pdf_filename']}")
    print("=" * 60)
    print("请用 LLM（claude/gpt）完成以下步骤：")
    print("  1. 读取每个 PDF 的前两页，提取作者单位（affiliations）")
    print("  2. 对每篇论文的 abstract 生成 150 字以内的中文总结（summary_cn）")
    print("  3. 将结果更新回 papers_record.xlsx 对应行")
    print("  4. 重建 viewer/papers_data.json")
    print("  5. 构建飞书 Markdown 消息并输出（cronjob 自动投递到飞书）")
    print("=" * 60)


if __name__ == "__main__":
    main()
